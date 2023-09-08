from openpyxl import Workbook
import openpyxl
import os
import re

os.getcwd()
wb = openpyxl.load_workbook('ParameterMainSheet.xlsx')
print(wb.sheetnames)
di = wb['data_input']


# 중간저장1
print(di['E5'].value, di['G5'].value, di['I5'].value)
print(di.cell(row=5,column=5).value)
print(str(di['E5'].value))

for row_index in range(5,32):
    contents_list = [di.cell(row=row_index,column=5).value,
                     str(di.cell(row=row_index,column=7).value or ''),
                     di.cell(row=row_index,column=9).value]
    contents_result = ' '.join(contents_list)
    di.cell(row=row_index,column=4).value = contents_result

file_name = 'temporary_result_check01.xlsx'
wb.save(filename=file_name)


# Letter Symbol Calculation
test_string = 'M04G03P02SC4.5D02E03'
a = re.sub(r'[0]+','',test_string)
b = re.findall('\d\.\d|\d',a)
print(a)
print(b)
print(b[1])

d = 0
for i in range(len(b)):
    d += float(b[i])
print(d)


# 중간저장2
wb = openpyxl.load_workbook('ParameterMainSheet.xlsx')
di = wb['data_input']
ds = wb['simple_work']

for row_index1 in range(5,32):
    for row_index2 in range(2,65):
        if di.cell(row=row_index1,column=9).value == ds.cell(row=row_index2,column=1).value:
            if di.cell(row=row_index1,column=10).value == ds.cell(row=row_index2,column=2).value:
                di.cell(row=row_index1,column=12).value = ds.cell(row=row_index2,column=7).value                
                
                symbol = ds.cell(row=row_index2,column=7).value
                symbol_trim = re.sub(r'[0]+','',symbol)
                symbol_split_list = re.findall('\d\.\d|\d',symbol_trim)
                mod = 0
                for i in range(len(symbol_split_list)):
                    mod += round(float(symbol_split_list[i]),1)
                di.cell(row=row_index1,column=13).value = mod
                
file_name = 'temporary_result_check02.xlsx'
wb.save(filename=file_name)


# Final 저장
wb = openpyxl.load_workbook('ParameterMainSheet.xlsx')
di = wb['data_input']
ds = wb['simple_work']

for row_index in range(5,32):
    contents_list = [di.cell(row=row_index,column=5).value,
                     str(di.cell(row=row_index,column=7).value or ''),
                     di.cell(row=row_index,column=9).value]
    contents_result = ' '.join(contents_list)
    di.cell(row=row_index,column=4).value = contents_result

for row_index1 in range(5,32):
    for row_index2 in range(2,65):
        if di.cell(row=row_index1,column=9).value == ds.cell(row=row_index2,column=1).value:
            if di.cell(row=row_index1,column=10).value == ds.cell(row=row_index2,column=2).value:
                di.cell(row=row_index1,column=12).value = ds.cell(row=row_index2,column=7).value                
                
                symbol = ds.cell(row=row_index2,column=7).value
                symbol_trim = re.sub(r'[0]+','',symbol)
                symbol_split_list = re.findall('\d\.\d|\d',symbol_trim)
                mod = 0
                for i in range(len(symbol_split_list)):
                    mod += round(float(symbol_split_list[i]),1)
                di.cell(row=row_index1,column=13).value = mod

dr = wb['result']
for row_index1 in range(5,32):
    dr.cell(row=row_index1, column=4).value = di.cell(row=row_index1,column=4).value   # contents 복사
    dr.cell(row=row_index1, column=7).value = di.cell(row=row_index1,column=11).value  # amount 복사
    dr.cell(row=row_index1, column=8).value = di.cell(row=row_index1,column=12).value  # symbol 복사
    dr.cell(row=row_index1, column=9).value = di.cell(row=row_index1,column=13).value  # mod 복사

file_name = 'FinalParameterMainSheet.xlsx'
wb.save(filename=file_name)

